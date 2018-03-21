from Tkinter import *
import tkFont
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import os

### Emulate the terminal in the GUI [BROKEN]
#import subprocess as sub
#p = sub.Popen('./home/pi/GEEN3400/GUI_Combo2_py2',stdout=sub.PIPE,stderr=sub.PIPE)
#output, errors = p.communicate()

### GUI definitions ###
root = Tk()
root.title("ElSA")
myFont = tkFont.Font(family = 'Helvetica', size = 12, weight = "bold")

leftFrame = Frame(root)
leftFrame.pack(side = LEFT)
rightFrame = Frame(root)
rightFrame.pack(side = RIGHT)
midFrame = Frame(root)
midFrame.pack(side = RIGHT)

### Variable definitions for widgts ###
checkVal1 = IntVar()
checkVal2 = IntVar()
sliderVal = DoubleVar()
rad = IntVar()

### Event Functions ###
def buttonPress():
    print('Slider value is {}' .format(sliderVal.get()))
def checkToggle():
    print('\nCheckbox #1  #2\n        {}  {}' .format(checkVal1.get(), checkVal2.get()))
def checkRadio():
    select = "Radio button " + str(rad.get()) + " selected."
    label.config(text = select)
def close():
    root.destroy() #close GUI
def mhello():
    mtext = ment.get()
    mlabel2 = Label(root, text = mtext).pack()
def pandas():
    ### PANDAS
    # Import desired inventory file and save as a variable
    file = 'Grocery.xlsx'
    xl = pd.ExcelFile(file)

    # Display all sheets that exist in spreadsheet file
    print(xl.sheet_names)

    # Generate DataFrame from imported spreadsheet and print
    df1 = xl.parse('Page 1')
    print (df1)
    shoplist = ["" for x in range(1)]

    item = "Start"
    while item != "Done":
        item = raw_input("Add item or type Done: ")
            #except ValueError:
                #print("Sorry, invalid input. Try adding another")
        if item != "Done":
            if df1['Item'].str.match(item).any():
                for j in range(len(shoplist)):
                    if shoplist[j] == "":
                        shoplist[j] = item
                        shoplist.append("")
                        print (shoplist)
                            # Use to test location of specific item-->> print(df1.loc[df1['Item'] == 'Nuts'])

                        # Find row in which desired item exists
                        for p in range(len(shoplist)-1):
                         #  rowselect = (df1.index[df1['Item'] == shoplist[p]])
                            rowselect = (df1.index[df1['Item'].str.match(shoplist[p])])
                            row4search = rowselect+1 #Variable used to search for numerical information in sheet
                            #Display int64 values
                            #print(rowselect)
                            #print(row4search)
                            #Display parsed row of item info
                            #print(df1.loc[rowselect])

                            wb = load_workbook('Grocery.xlsx') #Loads spreadsheet
                            sheet_1 = wb['Page 1'] #Extracts desired sheet
                            code = np.zeros(sheet_1.max_row) #Create blank array based on size of inventory list
                            cost = np.zeros(sheet_1.max_row)
                            aisle = np.zeros(sheet_1.max_row)
                            bay = np.zeros(sheet_1.max_row)
                     
                            for i in range(1,sheet_1.max_row):
                                code[i]=sheet_1.cell(row=i+1, column=2).value #Fill in data to blank arrays from spreadsheet
                                cost[i]=sheet_1.cell(row=i+1, column=3).value
                                aisle[i]=sheet_1.cell(row=i+1, column=4).value
                                bay[i]=sheet_1.cell(row=i+1, column=5).value

                        # Full info display
                            print (shoplist[p])
                            print ('Barcode', code[row4search]) #Display desired item barcode
                            print ('Price', cost[row4search]) #Display desired item cost
                            print ('Aisle', aisle[row4search]) #Display desired item isle
                            print ('Bay', bay[row4search]) #Display desired item bay
                        break
            else:
                print("Item not found. Please add a new Item.")
        else:
            break
menubar = Menu(root)

### Widgets ###
# print terminal into GUI [BROKEN]
#termf = Frame(root, height=40, width=50)
#termf.pack(fill=BOTH, expand=YES)
#wid = termf.winfo_id()
#os.system('xterm -into %d -geometry 40x20 -sb &' % wid)


# Text Display and Scroll Bar [BROKEN]
    #https://www.python-course.eu/tkinter_text_widget.php
#S = Scrollbar(root)
#T = Text(root, height=5, width=25)
#S.pack(side=RIGHT, fill=Y)
#T.pack(side=LEFT, fill=Y)
#S.config(command=T.yview)
#T.config(yscrollcommand=S.set)
#quote = "Something something something..."
#T.insert(END, quote)

# Text Box
T = Text(root, height=2, width=30)
T.pack()
T.insert(END, "Just a text Widget\nin two lines\n")

#Entry Box
ment = StringVar()
mlabel = Label(root, text='My label').pack()
mbutton = Button(root, text = 'Search', command = mhello, fg = 'red', bg = 'blue').pack()
mEntry = Entry(root, textvariable = ment).pack()

# Button - triggers a command when it is pressed
myButton = Button(midFrame, text = 'load', font = myFont, command = buttonPress, bg = 'bisque2', height = 1)
    #widgetName = widgetType(frame, text, font, what it does, base color, size)
myButton.pack(side = BOTTOM)
    #packs the given widget into the specified frame with all the other widgets in that frame with a certain location if given eg. BOTTOM

# 2 checkboxes - update attached variables and call command when toggled
check1 = Checkbutton(leftFrame, text = 'Lights ', variable = checkVal1, command = checkToggle)
check1.pack() 
check2 = Checkbutton(leftFrame, text = 'Sprinkler ', variable = checkVal2, command = checkToggle)
check2.pack()

# Radio buttons - good for mutually exclusive options
label = Label(leftFrame)
label.pack()
label.config(text = "No option selected")
R1 = Radiobutton(leftFrame, text = 'Option 1', variable = rad, value = 1, command = checkRadio)
R1.pack() 
R2 = Radiobutton(leftFrame, text = 'Option 2', variable = rad, value = 2, command = checkRadio)
R2.pack()

# Slider Bar - good for continuous parameter adjustment
slider = Scale(midFrame, variable = sliderVal, from_=100.0, to=0.0) #from is the top of the scale!
slider.pack()

# Spinbox - selectfrom a range of predetermined values
numOption = Spinbox(rightFrame, from_=1, to=10, width = 5)
numOption.pack(side = TOP)

# Exit button - closes the GUI
exitButton = Button(rightFrame, text = 'Exit', font = myFont, command = close, bg = 'red', width = 6)
exitButton.pack(side = BOTTOM)

# Create 'File', 'GEEN3400', and 'Help' menubars and submenus
filemenu = Menu(menubar, tearoff  = 0) #tearoff determines whether the drop down menu can be pulled away
filemenu.add_command(label = "Open", command = pandas) #opens a new window called placeholder
filemenu.add_command(label = "Save", command = pandas)
filemenu.add_command(label = "Save as...", command = pandas)
filemenu.add_separator() #adds a separating line between the command above and the command below
filemenu.add_command(label = "Quit", command = close)
menubar.add_cascade(label = "File", menu = filemenu)
    # The order of the add_cascade functions determines tehe order of the menu options
    
GEENmenu = Menu(menubar, tearoff = 0)
GEENmenu.add_command(label = "Amelia is cool", command = pandas)
GEENmenu.add_command(label = "Christian is cool", command = pandas)
GEENmenu.add_command(label = "Haley is cool", command = pandas)
GEENmenu.add_command(label = "Jared is cool", command = pandas)
menubar.add_cascade(label = "GEEN 3400", menu = GEENmenu)

helpmenu = Menu(menubar, tearoff = 0)
helpmenu.add_command(label = "How To", command = pandas)
helpmenu.add_separator()
helpmenu.add_command(label = "About", command = pandas)
menubar.add_cascade(label = "Help", menu = helpmenu)

#Display the menu
root.config(menu = menubar)
root.mainloop() #loops forever
