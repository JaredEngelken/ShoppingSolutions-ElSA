import Tkinter as tk
from Tkinter import *
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

class MainButton(object):

    def __init__(self):

        self.startbutton = Button(text = 'Cheesy Goodness', command = self.bringkeyboard)
        self.startbutton.pack()




        wb = load_workbook('Grocery.xlsx')
        self.sheet_1 = wb['Page 1']
        self.itemlist = [""]*(self.sheet_1.max_row)
        self.pricelist = [""]*(self.sheet_1.max_row)
        for t in range(0, self.sheet_1.max_row):
            self.itemlist[t] = self.sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1
            self.itemlist2 = ['{} '.format(elem) for elem in self.itemlist] #Formats to remove 'u' type from string

            self.pricelist[t] = ('   |${}|'.format(self.sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3
            self.pricelist2 = ['{} '.format(elem) for elem in self.pricelist] #Formats to remove 'u' type from string
            
        test_list = map(str.__add__, self.itemlist2, self.pricelist2)
        self.test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #Combines item with price info for user
        print self.test_list3

    def KEYBOARD(self):

        self.top = Toplevel()
        ws = self.top.winfo_screenwidth()
        print ws # width of the screen
        hs = self.top.winfo_screenheight() # height of the screen
        w = ws

        butheight = w/250
        butwidth = w/200
        h = 80*butheight
        # calculate x and y coordinates for the Tk self.top window
        x = (ws/2) - (w/2)
        y = (hs/2) - (h/2)
        # set the dimensions of the screen 
        # and where it is placed
        self.top.geometry('%dx%d+%d+%d' % (w, h, x, y))
        self.top.resizable(width=False, height=False)

        ##def updateSearch(searchstring):
        ##    search = Button(text = ''.join(searchstring))
        ##    search.grid(row = 0, column = 0)
            

        def AllCaps():

            #______All Caps_______________________________#
            #First Row#
            Qbut = Button(self.row1, text = 'Q', bg = 'white', command = lambda j='Q': Caps2Lower(j))
            Qbut.grid(column = 1, row = 1)
            Qbut.config(height = butheight, width = butwidth)

            Wbut = Button(self.row1, text = 'W', bg = 'white', command = lambda j='W': Caps2Lower(j))
            Wbut.grid(column = 2, row = 1)
            Wbut.config(height = butheight, width = butwidth)

            Ebut = Button(self.row1, text = 'E', bg = 'white', command = lambda j='E': Caps2Lower(j))
            Ebut.grid(column = 3, row = 1)
            Ebut.config(height = butheight, width = butwidth)

            Rbut = Button(self.row1, text = 'R', bg = 'white', command = lambda j='R': Caps2Lower(j))
            Rbut.grid(column = 4, row = 1)
            Rbut.config(height = butheight, width = butwidth)

            Tbut = Button(self.row1, text = 'T', bg = 'white', command = lambda j='T': Caps2Lower(j))
            Tbut.grid(column = 5, row = 1)
            Tbut.config(height = butheight, width = butwidth)

            Ybut = Button(self.row1, text = 'Y', bg = 'white', command = lambda j='Y': Caps2Lower(j))
            Ybut.grid(column = 6, row = 1)
            Ybut.config(height = butheight, width = butwidth)

            Ubut = Button(self.row1, text = 'U', bg = 'white', command = lambda j='U': Caps2Lower(j))
            Ubut.grid(column = 7, row = 1)
            Ubut.config(height = butheight, width = butwidth)

            Ibut = Button(self.row1, text = 'I', bg = 'white', command = lambda j='I': Caps2Lower(j))
            Ibut.grid(column = 8, row = 1)
            Ibut.config(height = butheight, width = butwidth)

            Obut = Button(self.row1, text = 'O', bg = 'white', command = lambda j='O': Caps2Lower(j))
            Obut.grid(column = 9, row = 1)
            Obut.config(height = butheight, width = butwidth)

            Pbut = Button(self.row1, text = 'P', bg = 'white', command = lambda j='P': Caps2Lower(j))
            Pbut.grid(column = 10, row = 1)
            Pbut.config(height = butheight, width = butwidth)

            #IF PADDING NEEDED --> Obut.grid(column = 9, row = 1, padx = w/40, pady = h/40)

            #Second Row#
            Abut = Button(self.row2, text = 'A', bg = 'white', command = lambda j='A': Caps2Lower(j))
            Abut.grid(column = 1, row = 1)
            Abut.config(height = butheight, width = butwidth)

            Sbut = Button(self.row2, text = 'S', bg = 'white', command = lambda j='S': Caps2Lower(j))
            Sbut.grid(column = 2, row = 1)
            Sbut.config(height = butheight, width = butwidth)

            Dbut = Button(self.row2, text = 'D', bg = 'white', command = lambda j='D': Caps2Lower(j))
            Dbut.grid(column = 3, row = 1)
            Dbut.config(height = butheight, width = butwidth)

            Fbut = Button(self.row2, text = 'F', bg = 'white', command = lambda j='F': Caps2Lower(j))
            Fbut.grid(column = 4, row = 1)
            Fbut.config(height = butheight, width = butwidth)

            Gbut = Button(self.row2, text = 'G', bg = 'white', command = lambda j='G': Caps2Lower(j))
            Gbut.grid(column = 5, row = 1)
            Gbut.config(height = butheight, width = butwidth)

            Hbut = Button(self.row2, text = 'H', bg = 'white', command = lambda j='H': Caps2Lower(j))
            Hbut.grid(column = 6, row = 1)
            Hbut.config(height = butheight, width = butwidth)

            Jbut = Button(self.row2, text = 'J', bg = 'white', command = lambda j='J': Caps2Lower(j))
            Jbut.grid(column = 7, row = 1)
            Jbut.config(height = butheight, width = butwidth)

            Kbut = Button(self.row2, text = 'K', bg = 'white', command = lambda j='K': Caps2Lower(j))
            Kbut.grid(column = 8, row = 1)
            Kbut.config(height = butheight, width = butwidth)

            Lbut = Button(self.row2, text = 'L', bg = 'white', command = lambda j='L': Caps2Lower(j))
            Lbut.grid(column = 9, row = 1)
            Lbut.config(height = butheight, width = butwidth)

            #Third Row#
            ShiftBut = Button(self.row3, text = 'Shift', bg = 'white', command = LowerCase)
            ShiftBut.grid(column = 1, row = 1)
            ShiftBut.config(height = butheight, width = butwidth*2)

            Zbut = Button(self.row3, text = 'Z', bg = 'white', command = lambda j='Z': Caps2Lower(j))
            Zbut.grid(column = 3, row = 1)
            Zbut.config(height = butheight, width = butwidth)

            Xbut = Button(self.row3, text = 'X', bg = 'white', command = lambda j='X': Caps2Lower(j))
            Xbut.grid(column = 4, row = 1)
            Xbut.config(height = butheight, width = butwidth)

            Cbut = Button(self.row3, text = 'C', bg = 'white', command = lambda j='C': Caps2Lower(j))
            Cbut.grid(column = 5, row = 1)
            Cbut.config(height = butheight, width = butwidth)

            Vbut = Button(self.row3, text = 'V', bg = 'white', command = lambda j='V': Caps2Lower(j))
            Vbut.grid(column = 6, row = 1)
            Vbut.config(height = butheight, width = butwidth)

            Bbut = Button(self.row3, text = 'B', bg = 'white', command = lambda j='B': Caps2Lower(j))
            Bbut.grid(column = 7, row = 1)
            Bbut.config(height = butheight, width = butwidth)

            Nbut = Button(self.row3, text = 'N', bg = 'white', command = lambda j='N': Caps2Lower(j))
            Nbut.grid(column = 8, row = 1)
            Nbut.config(height = butheight, width = butwidth)

            Mbut = Button(self.row3, text = 'M', bg = 'white', command = lambda j='M': Caps2Lower(j))
            Mbut.grid(column = 9, row = 1)
            Mbut.config(height = butheight, width = butwidth)

            DelBut = Button(self.row3, text = 'Delete', bg = 'white', command = delete)
            DelBut.grid(column = 10, row = 1)
            DelBut.config(height = butheight, width = butwidth*2)


            #Fourth Row#
            NumBut = Button(self.row4, text = '123', bg = 'white')
            NumBut.grid(column = 1, row = 1)
            NumBut.config(height = butheight, width = butwidth*2)

            Spacebut = Button(self.row4, text = 'Space', bg = 'white',command=lambda:action(' '))
            Spacebut.grid(column = 3, row = 1)
            Spacebut.config(height = butheight, width = butwidth*8)

            Clearbut = Button(self.row4, text = 'Clear Search', bg = 'white',command = clearsearch)
            Clearbut.grid(column = 4, row = 1)
            Clearbut.config(height = butheight, width = butwidth*2)

        def LowerCase():
            #self = Frame(self.top)

            #______All Caps_______________________________#
            #First Row#
            Qbut = Button(self.row1, text = 'q', bg = 'white',command=lambda:action('q'))
            Qbut.grid(column = 1, row = 1)
            Qbut.config(height = butheight, width = butwidth)

            Wbut = Button(self.row1, text = 'w', bg = 'white',command=lambda:action('w'))
            Wbut.grid(column = 2, row = 1)
            Wbut.config(height = butheight, width = butwidth)

            Ebut = Button(self.row1, text = 'e', bg = 'white',command=lambda:action('e'))
            Ebut.grid(column = 3, row = 1)
            Ebut.config(height = butheight, width = butwidth)

            Rbut = Button(self.row1, text = 'r', bg = 'white',command=lambda:action('r'))
            Rbut.grid(column = 4, row = 1)
            Rbut.config(height = butheight, width = butwidth)

            Tbut = Button(self.row1, text = 't', bg = 'white',command=lambda:action('t'))
            Tbut.grid(column = 5, row = 1)
            Tbut.config(height = butheight, width = butwidth)

            Ybut = Button(self.row1, text = 'y', bg = 'white',command=lambda:action('y'))
            Ybut.grid(column = 6, row = 1)
            Ybut.config(height = butheight, width = butwidth)

            Ubut = Button(self.row1, text = 'u', bg = 'white',command=lambda:action('u'))
            Ubut.grid(column = 7, row = 1)
            Ubut.config(height = butheight, width = butwidth)

            Ibut = Button(self.row1, text = 'i', bg = 'white',command=lambda:action('i'))
            Ibut.grid(column = 8, row = 1)
            Ibut.config(height = butheight, width = butwidth)

            Obut = Button(self.row1, text = 'o', bg = 'white',command=lambda:action('o'))
            Obut.grid(column = 9, row = 1)
            Obut.config(height = butheight, width = butwidth)

            Pbut = Button(self.row1, text = 'p', bg = 'white',command=lambda:action('p'))
            Pbut.grid(column = 10, row = 1)
            Pbut.config(height = butheight, width = butwidth)


            #Second Row#
            Abut = Button(self.row2, text = 'a', bg = 'white',command=lambda:action('a'))
            Abut.grid(column = 1, row = 1)
            Abut.config(height = butheight, width = butwidth)

            Sbut = Button(self.row2, text = 's', bg = 'white',command=lambda:action('s'))
            Sbut.grid(column = 2, row = 1)
            Sbut.config(height = butheight, width = butwidth)

            Dbut = Button(self.row2, text = 'd', bg = 'white',command=lambda:action('d'))
            Dbut.grid(column = 3, row = 1)
            Dbut.config(height = butheight, width = butwidth)

            Fbut = Button(self.row2, text = 'f', bg = 'white',command=lambda:action('f'))
            Fbut.grid(column = 4, row = 1)
            Fbut.config(height = butheight, width = butwidth)

            Gbut = Button(self.row2, text = 'g', bg = 'white',command=lambda:action('g'))
            Gbut.grid(column = 5, row = 1)
            Gbut.config(height = butheight, width = butwidth)

            Hbut = Button(self.row2, text = 'h', bg = 'white',command=lambda:action('h'))
            Hbut.grid(column = 6, row = 1)
            Hbut.config(height = butheight, width = butwidth)

            Jbut = Button(self.row2, text = 'j', bg = 'white',command=lambda:action('j'))
            Jbut.grid(column = 7, row = 1)
            Jbut.config(height = butheight, width = butwidth)

            Kbut = Button(self.row2, text = 'k', bg = 'white',command=lambda:action('k'))
            Kbut.grid(column = 8, row = 1)
            Kbut.config(height = butheight, width = butwidth)

            Lbut = Button(self.row2, text = 'l', bg = 'white',command=lambda:action('l'))
            Lbut.grid(column = 9, row = 1)
            Lbut.config(height = butheight, width = butwidth)

            #Third Row#
            ShiftBut = Button(self.row3, text = 'Shift', bg = 'white', command = AllCaps)
            ShiftBut.grid(column = 1, row = 1)
            ShiftBut.config(height = butheight, width = butwidth*2)

            Zbut = Button(self.row3, text = 'z', bg = 'white',command=lambda:action('z'))
            Zbut.grid(column = 3, row = 1)
            Zbut.config(height = butheight, width = butwidth)

            Xbut = Button(self.row3, text = 'x', bg = 'white',command=lambda:action('x'))
            Xbut.grid(column = 4, row = 1)
            Xbut.config(height = butheight, width = butwidth)

            Cbut = Button(self.row3, text = 'c', bg = 'white',command=lambda:action('c'))
            Cbut.grid(column = 5, row = 1)
            Cbut.config(height = butheight, width = butwidth)

            Vbut = Button(self.row3, text = 'v', bg = 'white',command=lambda:action('v'))
            Vbut.grid(column = 6, row = 1)
            Vbut.config(height = butheight, width = butwidth)

            Bbut = Button(self.row3, text = 'b', bg = 'white',command=lambda:action('b'))
            Bbut.grid(column = 7, row = 1)
            Bbut.config(height = butheight, width = butwidth)

            Nbut = Button(self.row3, text = 'n', bg = 'white',command=lambda:action('n'))
            Nbut.grid(column = 8, row = 1)
            Nbut.config(height = butheight, width = butwidth)

            Mbut = Button(self.row3, text = 'm', bg = 'white',command=lambda:action('m'))
            Mbut.grid(column = 9, row = 1)
            Mbut.config(height = butheight, width = butwidth)

            DelBut = Button(self.row3, text = 'Delete', bg = 'white', command = delete)
            DelBut.grid(column = 10, row = 1)
            DelBut.config(height = butheight, width = butwidth*2)


            #Fourth Row#
            NumBut = Button(self.row4, text = '123', bg = 'white')
            NumBut.grid(column = 1, row = 1)
            NumBut.config(height = butheight, width = butwidth*2)

            Spacebut = Button(self.row4, text = 'Space', bg = 'white',command=lambda:action(' '))
            Spacebut.grid(column = 3, row = 1)
            Spacebut.config(height = butheight, width = butwidth*8)

            Clearbut = Button(self.row4, text = 'Clear Search', bg = 'white', command = clearsearch)
            Clearbut.grid(column = 4, row = 1)
            Clearbut.config(height = butheight, width = butwidth*2)

        #All Called Functions
        def Caps2Lower(WhichBut):
            action(WhichBut)
            LowerCase()
            
        def delete():
            self.txt = self.e.get()[:-1]
            self.e.delete(0,END)
            self.e.insert(0,self.txt)
            if len(self.txt) == 0:
                self.data = []
                self.listbox.configure(height = 0)
                self.listbox_update(self)

                AllCaps()
            value = self.e.get().strip().lower()
            if len(self.txt) != 0:
                self.data = []
                for item in self.test_list3:
                    if value in item.lower():
                        self.data.append(item)
                self.listbox_update(self)
                
        def clearsearch():
            self.e.delete(0,END)
            self.data = []
            self.listbox_update(self)
            AllCaps()

        def action(argi): 
            """pressed button's value is inserted into the end of the text area"""
            self.e.insert(END,argi)
            value = self.e.get()
            value = value.strip().lower()
            # get self.data from test_list
            if value == '':
                self.data = []
            else:
                self.data = []
                for item in self.test_list3:
                    if value in item.lower():
                        self.data.append(item)                
            # update self.data in listbox
            self.listbox_update(self)
            
    #^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^#

        
        #_______________________________Main_Loop_________________________________#

        self.row0 = Frame(self.top)
        self.row1 = Frame(self.top)
        self.row2 = Frame(self.top)
        self.row3 = Frame(self.top)
        self.row4 = Frame(self.top)

        self.row0.grid(row=0, padx = w/64)
        self.row0.rowconfigure(0, weight = 1)
        self.row1.grid(row=1, padx = w/64)
        self.row1.rowconfigure(1, weight = 1)
        self.row2.grid(row=2, padx = w/64)
        self.row2.rowconfigure(2, weight = 1)
        self.row3.grid(row=3, padx = w/64)
        self.row3.rowconfigure(3, weight = 1)
        self.row4.grid(row=4, padx = w/64)
        self.row4.rowconfigure(4, weight = 1)


        #Numbers#
        Percentbut = Button(self.row0, text = '%', bg = 'white',command=lambda:action('%'))
        Percentbut.grid(column = 1, row = 1)
        Percentbut.config(height = butheight/2, width = butwidth)

        Onebut = Button(self.row0, text = '1', bg = 'white',command=lambda:action('1'))
        Onebut.grid(column = 2, row = 1)
        Onebut.config(height = butheight/2, width = butwidth)

        Twobut = Button(self.row0, text = '2', bg = 'white',command=lambda:action('2'))
        Twobut.grid(column = 3, row = 1)
        Twobut.config(height = butheight/2, width = butwidth)

        Threebut = Button(self.row0, text = '3', bg = 'white',command=lambda:action('3'))
        Threebut.grid(column = 4, row = 1)
        Threebut.config(height = butheight/2, width = butwidth)

        Fourbut = Button(self.row0, text = '4', bg = 'white',command=lambda:action('4'))
        Fourbut.grid(column = 5, row = 1)
        Fourbut.config(height = butheight/2, width = butwidth)

        Fivebut = Button(self.row0, text = '5', bg = 'white',command=lambda:action('5'))
        Fivebut.grid(column = 6, row = 1)
        Fivebut.config(height = butheight/2, width = butwidth)

        Sixbut = Button(self.row0, text = '6', bg = 'white',command=lambda:action('6'))
        Sixbut.grid(column = 7, row = 1)
        Sixbut.config(height = butheight/2, width = butwidth)

        Sevenbut = Button(self.row0, text = '7', bg = 'white',command=lambda:action('7'))
        Sevenbut.grid(column = 8, row = 1)
        Sevenbut.config(height = butheight/2, width = butwidth)

        Eightbut = Button(self.row0, text = '8', bg = 'white',command=lambda:action('8'))
        Eightbut.grid(column = 9, row = 1)
        Eightbut.config(height = butheight/2, width = butwidth)

        Ninebut = Button(self.row0, text = '9', bg = 'white',command=lambda:action('9'))
        Ninebut.grid(column = 10, row = 1)
        Ninebut.config(height = butheight/2, width = butwidth)

        Zerobut = Button(self.row0, text = '0', bg = 'white',command=lambda:action('0'))
        Zerobut.grid(column = 11, row = 1)
        Zerobut.config(height = butheight/2, width = butwidth)



        #Start Program in AllCaps#
        AllCaps()
        self.top.mainloop()

    def listbox_update(data, self):
        # delete previous self.data
        self.listbox.delete(0, 'end')
        # sorting self.data 
        self.data = sorted(self.data, key=str.lower)
        # put new self.data
        for item in self.data:
            self.listbox.insert('end', item)

    def on_select(self, listbox):


    # Takes user selection and removes price information in order to use selected item to search for other info
    #______________________________________________________________________________________________________________#
        all_items = self.listbox.get(0, 'end')

        if len(all_items) != 0:
            list_item_number = self.listbox.curselection()
            print list_item_number
            print all_items
            grabitem = [all_items[item] for item in list_item_number]

            wb = load_workbook('Grocery.xlsx')
            self.sheet_1 = wb['Page 1']
            self.itemlist = [""]*(self.sheet_1.max_row)
            self.pricelist = [""]*(self.sheet_1.max_row)

            for t in range(0, self.sheet_1.max_row):
                self.itemlist[t] = self.sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1 describing item
                self.itemlist2 = ['{} '.format(elem) for elem in self.itemlist] #Formats to remove 'u' type from string

                self.pricelist[t] = ('   |${}|'.format(self.sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3 describing price
                self.pricelist2 = ['{} '.format(elem) for elem in self.pricelist] #Formats to remove 'u' type from string
            test_list = map(str.__add__, self.itemlist2, self.pricelist2) #combines lists into one in order to find index within list of selected item
            self.test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #clean up list
            print self.test_list3

            grabitem = str(grabitem[0])


            grabitemindex = self.test_list3.index(grabitem)+1 #finds index of selected item
            pricelength = len(self.pricelist2[grabitemindex]) #finds length of displayed price info in order to determine how long of a string to delete
            item = grabitem[:-pricelength] #deletes price info leaving only item title

        #______________________________________________________________________________________________________________#

        #Below runs item through Pandas
            file = 'Grocery.xlsx'
            xl = pd.ExcelFile(file)
            
            # Generate self.dataFrame from imported spreadsheet and print
            df1 = xl.parse('Page 1')
            print (df1)

            #regenerate list of items in inventory
            list1 = [""]*(self.sheet_1.max_row)
            for t in range(0, self.sheet_1.max_row):
                list1[t] = self.sheet_1.cell(row = t+1, column = 1).value
                list2 = filter(lambda a: a != "Item", list1)
                list3 = ['{} '.format(elem) for elem in list2]
            print list3
            
            rowselect = list3.index(item)
            row4search = rowselect+2 #find which row item exists in
            print row4search


            code=self.sheet_1.cell(row=row4search, column=2).value #Fill in self.data to blank arrays from spreadsheet
            cost=self.sheet_1.cell(row=row4search, column=3).value
            aisle=self.sheet_1.cell(row=row4search, column=4).value
            bay=self.sheet_1.cell(row=row4search, column=5).value

            # Full info display
            print (self.sheet_1.cell(row=row4search, column=1).value)
            print ('Barcode {}'.format(code)) #Display desired item barcode
            print ('Price ${}'.format(cost)) #Display desired item cost
            print ('Aisle {}'.format(aisle)) #Display desired item isle
            print ('Bay {}'.format(bay)) #Display desired item bay

        
    def bringkeyboard(self):

        self.startbutton.destroy()

        self.e = Entry()
        self.e.grid(row = 1, column = 1)
        self.e.focus()

        self.listbox = Listbox(font = 10)
        self.listbox.grid(row = 2, column = 1)
        self.listbox.configure(height = 0)
        self.listbox.bind('<<ListboxSelect>>', self.on_select)
        self.data = []
        self.listbox_update(self)
        
        self.cancelbutton = Button(text = 'Cancel')
        self.cancelbutton.grid(row = 1, column = 2)
        self.cancelbutton.configure(command = lambda: self.searchyboi(self.cancelbutton, self.e, self.listbox, self.top))

        go = self.KEYBOARD()

    def searchyboi(self, cancelbutton, e, listbox, top):

        self.e.destroy()
        self.listbox.destroy()
        self.cancelbutton.destroy()
        self.top.destroy()

        self.startbutton = Button(text = 'Search', command = self.bringkeyboard)
        self.startbutton.grid(row = 1, column = 1)



if __name__ == "__main__":
    master = tk.Tk()
    MainButton()
    master.mainloop()
        
#TO DO#
    #implement changes to listbox.bind and event attribute errors
        
