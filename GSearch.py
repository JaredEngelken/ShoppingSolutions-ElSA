import Tkinter as tk
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

#//////////////////////////////////////////////////////////////////////////////////////////////////////////////#
def on_keyrelease(event):
    # get text from entry
    value = event.widget.get()
    value = value.strip().lower()
    # get data from test_list
    if value == '':
        data = []
    else:
        data = []
        for item in test_list3:
            if value in item.lower():
                data.append(item)                
    # update data in listbox
    listbox_update(data)
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

#//////////////////////////////////////////////////////////////////////////////////////////////////////////////#
def listbox_update(data):
    # delete previous data
    listbox.delete(0, 'end')
    # sorting data 
    data = sorted(data, key=str.lower)
    # put new data
    for item in data:
        listbox.insert('end', item)
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
        
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////#
def on_select(event):
# Takes user selection and removes price information in order to use selected item to search for other info
#______________________________________________________________________________________________________________#
    grabitem = event.widget.get(event.widget.curselection()) #the initial selection of the user
    for t in range(0, sheet_1.max_row):
        itemlist[t] = sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1 describing item
        itemlist2 = ['{} '.format(elem) for elem in itemlist] #Formats to remove 'u' type from string

        pricelist[t] = ('   |${}|'.format(sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3 describing price
        pricelist2 = ['{} '.format(elem) for elem in pricelist] #Formats to remove 'u' type from string
    
    test_list = map(str.__add__, itemlist2, pricelist2) #combines lists into one in order to find index within list of selected item
    test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #clean up list
    print test_list3

    grabitemindex = test_list3.index(grabitem)+1 #finds index of selected item
    pricelength = len(pricelist2[grabitemindex]) #finds length of displayed price info in order to determine how long of a string to delete

    item = grabitem[:-pricelength] #deletes price info leaving only item title
#______________________________________________________________________________________________________________#

#Below runs item through Pandas
    file = 'Grocery.xlsx'
    xl = pd.ExcelFile(file)
    
    # Generate DataFrame from imported spreadsheet and print
    df1 = xl.parse('Page 1')
    print (df1)

    #regenerate list of items in inventory
    list1 = [""]*(sheet_1.max_row)
    for t in range(0, sheet_1.max_row):
        list1[t] = sheet_1.cell(row = t+1, column = 1).value
        list2 = filter(lambda a: a != "Item", list1)
        list3 = ['{} '.format(elem) for elem in list2]
    print list3
    
    rowselect = list3.index(item)
    row4search = rowselect+2 #find which row item exists in
    print row4search


    code=sheet_1.cell(row=row4search, column=2).value #Fill in data to blank arrays from spreadsheet
    cost=sheet_1.cell(row=row4search, column=3).value
    aisle=sheet_1.cell(row=row4search, column=4).value
    bay=sheet_1.cell(row=row4search, column=5).value

    # Full info display
    print (sheet_1.cell(row=row4search, column=1).value)
    print ('Barcode {}'.format(code)) #Display desired item barcode
    print ('Price ${}'.format(cost)) #Display desired item cost
    print ('Aisle {}'.format(aisle)) #Display desired item isle
    print ('Bay {}'.format(bay)) #Display desired item bay
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

# --- Main ---

#Loads workbook to look for items in inventory
wb = load_workbook('Grocery.xlsx')
sheet_1 = wb['Page 1']
itemlist = [""]*(sheet_1.max_row)
pricelist = [""]*(sheet_1.max_row)
for t in range(0, sheet_1.max_row):
    itemlist[t] = sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1
    itemlist2 = ['{} '.format(elem) for elem in itemlist] #Formats to remove 'u' type from string

    pricelist[t] = ('   |${}|'.format(sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3
    pricelist2 = ['{} '.format(elem) for elem in pricelist] #Formats to remove 'u' type from string
    
test_list = map(str.__add__, itemlist2, pricelist2)
test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #Combines item with price info for user
print test_list3

root = tk.Tk()

entry = tk.Entry(root)
entry.pack()
entry.bind('<KeyRelease>', on_keyrelease)

listbox = tk.Listbox(root)
listbox.pack()
listbox.bind('<<ListboxSelect>>', on_select)
listbox_update([])

root.mainloop()
