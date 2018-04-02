from Tkinter import *
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

master = Tk()

GFVar = IntVar()
Checkbutton(master, text = "Gluten Free",variable = GFVar).grid(row = 0,sticky = W)
VeganVar = IntVar()
Checkbutton(master, text = "Vegan", variable = VeganVar).grid(row = 1, sticky = W)
DFVar = IntVar()
Checkbutton(master, text = "Dairy Free", variable = DFVar).grid(row = 2,sticky = W)
VegVar = IntVar()
Checkbutton(master, text = "Vegetarian",variable = VegVar).grid(row = 3,sticky = W)


def finddatrecipe():
    diettypes = ['Gluten Free', 'Vegan','Dairy Free', 'Vegetarian','']

    if GFVar.get() == 1:
        diettypes[0] = 'Gluten Free'
        print diettypes
    else: 
        diettypes[0] = ''
        print diettypes
    if VeganVar.get() == 1:
        diettypes[1] = 'Vegan'
        print diettypes
    else:
        diettypes[1] = ''
        print diettypes
    if DFVar.get() == 1:
        diettypes[2] = 'Dairy Free'
        print diettypes
    else:
        diettypes[2] = ''
        print diettypes
    if VegVar.get() == 1:
        diettypes[3] = 'Vegetarian'
        print diettypes
    else:
        diettypes[3] = ''
        print diettypes

           
    dietrestrict = (filter(None, diettypes))
    dietrestrict.append('')
    print dietrestrict

    file = 'RecipeBook.xlsx'
    recipebook = pd.ExcelFile(file)

    print(recipebook.sheet_names)

    #for options in range(len(dietrestrict)):
    list1 = recipebook.parse('All')
    print list1

    finalset = ['Blank']
    spot = 0


    for h in range(len(dietrestrict)-1):
        if dietrestrict[h] == 'Gluten Free':
            glutenrow = (list1.index.values[list1['Gluten Free'].str.match('Yes')])+1
            print glutenrow
            finalset[spot] = glutenrow
            spot = spot+1
            finalset.append([])
            print finalset
        if dietrestrict[h] == 'Vegan':
            veganrow = (list1.index.values[list1['Vegan'].str.match('Yes')])+1
            print veganrow
            finalset[spot] = veganrow
            spot = spot+1
            finalset.append([])
            print finalset
        if dietrestrict[h] == 'Dairy Free':
            dairyrow = (list1.index.values[list1['Dairy Free'].str.match('Yes')])+1
            print dairyrow
            finalset[spot] = dairyrow
            spot = spot+1
            finalset.append([])
            print finalset
        if dietrestrict[h] == 'Vegetarian':
            vegetarian = (list1.index.values[list1['Vegetarian'].str.match('Yes')])+1
            print vegetarian
            finalset[spot] = vegetarian
            spot = spot+1
            finalset.append([])
            print finalset
    #Make as many of these if statements as there are dietary options#
    #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv#        
    if (len(dietrestrict)-1) == 1:
        searchset = finalset[0]
        #w = Message(master, text = '').grid(row = 3, column = 2, sticky = W)
        print searchset
    if (len(dietrestrict)-1) == 2:
        finalset = finalset[:-1]
        print finalset
        set1 = finalset[0]
        set2 = finalset[1]
        def intersection(set1, set2):
            return list(set(set1) & set(set2))
        searchset = (intersection(set1, set2))
        #w = Message(master, text = '').grid(row = 3, column = 2, sticky = W)
        print searchset
    if (len(dietrestrict)-1) == 3:
        finalset = finalset[:-1]
        print finalset
        set1 = finalset[0]
        set2 = finalset[1]
        set3 = finalset[2]
        def intersection(set1, set2, set3):
            return list(set(set1) & set(set2) & set(set3))
        searchset = (intersection(set1, set2, set3))
        #w = Message(master, text = '').grid(row = 3, column = 2, sticky = W)
        print searchset
    if (len(dietrestrict)-1) == 4:
        finalset = finalset[:-1]
        print finalset
        set1 = finalset[0]
        set2 = finalset[1]
        set3 = finalset[2]
        set4 = finalset[3]
        def intersection(set1, set2, set3, set4):
            return list(set(set1) & set(set2) & set(set3) & set(set4))
        searchset = (intersection(set1, set2, set3, set4))
        #w = Message(master, text = '').grid(row = 3, column = 2, sticky = W)
        print searchset

    #if len(searchset) == 0:
    #    mes = "Sorry. No recipes match those dietary restrictions."
    #    w = Message(master, text = mes).grid(row = 3, column = 2, sticky = W)
        

    #^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^#

    secondload = load_workbook('RecipeBook.xlsx')
    All = secondload['All']


    for result in range(len(searchset)):
        val = (searchset[result])+1
        givenRecipe = (All.cell(row = val, column = 5)).value #Make Column value equal to the number of dietary options plus 1
        print givenRecipe
        descriptor = [""]*(All.max_column)
        ingredients = [""]*(All.max_column)

        for thing in range(9, (All.max_column)+1): #Make starting range equal to number of dietary restrictions plus 5
            if ((All.cell(row = val, column = thing)).value) != []:
                descriptor[thing-8] = ((All.cell(row = val+1, column = thing)).value)
                desc1 = filter(lambda a: a != "", descriptor)
                desc2 = [str(item) for item in desc1]
                desc3 = filter(lambda b: b != "None", desc2)
                ingredients[thing-8] = (All.cell(row = val, column = thing)).value
                allIngr1 = filter(lambda a: a != "", ingredients)
                allIngr2 = [str(item) for item in allIngr1]
                allIngr3 = filter(lambda b: b != "None", allIngr2)
            else:
                break
        #print desc3
        #print allIngr3
        desc4 = ['{} '.format(elem) for elem in desc3]
        displayIngr = map(str.__add__,desc4, allIngr3)
        print displayIngr
########################################################################################
RecButton = Button(master, text = 'Find Me Recipes', command = finddatrecipe) 
RecButton.grid(row = 4, sticky = W, pady = 4)



mainloop()


