#import tkinter in a universal way that would work for both python2 and python3
try:
    #Python 2
    import Tkinter as tk
except ImportError:
    #Python 3
    import tkinter as tk
#from Tkinter import *
import tkFont
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import os
import time
#import keyboard

class GUI():
    def __init__(self, master):
        self.master = master

        self.myFont = tkFont.Font(family = 'Helvetica', size = 12, weight = "bold")

        self.topFrame1 = tk.Frame(self.master, bg='lightskyblue') #topFrame1 = tk.Frame(root, bg='lightskyblue', width = 1920, height = 200, padx = 5, pady = 5)
        self.topFrame2 = tk.Frame(self.master, bg='lightskyblue')
        self.topFrame3 = tk.Frame(self.master, bg='lightskyblue')
        self.leftFrame1 = tk.Frame(self.master, bg='blue')
        self.centerFrame1 = tk.Frame(self.master, bg='deepskyblue')
        self.rightFrame1 = tk.Frame(self.master, bg='darkblue')
        self.bottomFrame1 = tk.Frame(self.master, bg='magenta')

        self.leftFrame1.columnconfigure(0, weight = 1, uniform = 'same')
        self.centerFrame1.columnconfigure(1, weight = 2, uniform = 'same')
        self.rightFrame1.columnconfigure(2, weight = 1, uniform = 'same')
        self.master.grid_rowconfigure(0, weight = 1, uniform = 'same1')
        self.master.grid_rowconfigure(1, weight = 7, uniform = 'same1')
        self.master.grid_rowconfigure(2, weight = 2, uniform = 'same1')
        self.master.grid_columnconfigure(0, weight = 1, uniform = 'same2')
        self.master.grid_columnconfigure(1, weight = 1, uniform = 'same2')
        self.master.grid_columnconfigure(2, weight = 1, uniform = 'same2')
        self.topFrame1.grid_rowconfigure(0, weight = 1, uniform = 'same3')
        self.topFrame2.grid_rowconfigure(1, weight = 1, uniform = 'same3')
        self.topFrame3.grid_rowconfigure(2, weight = 1, uniform = 'same3')

        self.topFrame1.grid(row=0, column = 0, sticky="nsew")
        self.topFrame2.grid(row=0, column = 1, sticky="nsew")
        self.topFrame3.grid(row=0, column = 2, sticky="nsew")
        self.leftFrame1.grid(row=1, column = 0, sticky="nsew")
        self.centerFrame1.grid(row=1, column = 1, sticky="nsew")
        self.rightFrame1.grid(row=1, column = 2, sticky="nsew")
        self.bottomFrame1.grid(row=2, columnspan = 3, sticky="nsew")

        self.itemMaster = ['']
        self.priceMaster = ['']

        self.bottomFrameWidth = self.bottomFrame1.winfo_width()

        # Shopping Cart Label
        self.shoppingCartLabel = tk.Label(self.rightFrame1, text = "Your Shopping Cart:", bg = 'orchid', fg = 'floralwhite', font = self.myFont)
        self.shoppingCartLabel.pack(fill = tk.X)

        # Total Cost Label
        self.totalCostLabel = tk.Label(self.rightFrame1, text = "Total Cost: ", bg = 'cornflowerblue', fg = 'lavenderblush', font = self.myFont)
        self.totalCostLabel.pack(fill = tk.X, side = tk.BOTTOM)

        # Shopping List Label
        self.addCartLabel = tk.Label(self.leftFrame1, text = "Shopping List\nClick to add items to cart: ", bg = 'cornflowerblue', fg = 'lavenderblush', font = self.myFont)
        self.addCartLabel.pack(fill = tk.X)
        
        # Search button
        self.searchButton = tk.Button(self.centerFrame1, text = 'Search', bg = 'mediumslateblue', fg = 'snow', font = self.myFont, command = self.gsearch, height = 9)
        self.searchButton.pack(fill = tk.X)

        # Store Logo
        self.logoLabel = tk.Label(self.topFrame2, text = 'Place logo here!', font = self.myFont, height = 10, width = 30)
        self.logoLabel.pack()

        # Ad Button
        self.adButton = tk.Button(self.bottomFrame1, text = 'Place ad here!', font = self.myFont, height = 96, width = 800)
        self.photoMilk = tk.PhotoImage(file = "HorizonAd.png")
        self.adButton.config(image = self.photoMilk)
        self.adButton.pack()

        # Recipe Builder Button
        self.recipeButton = tk.Button(self.centerFrame1, text = 'Place recipes here!', font = self.myFont, height = 10)
        self.recipeButton.pack(side = tk.BOTTOM, fill = tk.X)

        #self.test_list = test_list
        #self.test_list3 = test_list3


    def close(self):
        self.master.destroy() #close GUI

    def on_select(self, event):
    # Takes user selection and removes price information in order to use selected item to search for other info
    #______________________________________________________________________________________________________________#
        grabitem = event.widget.get(event.widget.curselection()) #the initial selection of the user
        for t in range(0, sheet_1.max_row):
            itemlist[t] = sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1 describing item
            itemlist2 = ['{} '.format(elem) for elem in itemlist] #Formats to remove 'u' type from string

            pricelist[t] = ('   |${}|'.format(sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3 describing price
            pricelist2 = ['{} '.format(elem) for elem in pricelist] #Formats to remove 'u' type from string
        
        test_list = map(str.__add__, itemlist2, pricelist2) #combines lists into one in order to find index within list of selected item
        self.test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #clean up list
        print test_list3

        grabitemindex = test_list3.index(grabitem)+1 #finds index of selected item
        pricelength = len(pricelist2[grabitemindex]) #finds length of displayed price info in order to determine how long of a string to delete

        item = grabitem[:-pricelength] #deletes price info leaving only item title
        
##        for d in range(0, k):
##            if itemMaster[d] == '':
##                itemMaster[d] == item
##                itemMaster.append('')
##                print itemMaster
##                break
##
##        for p in range(len(priceMaster)):
##            if priceMaster[p] == '':
##                priceMaster[p] == item
##                priceMaster.append('')
##                print priceMaster
##                break
        
    #______________________________________________________________________________________________________________#

    #Below runs item through Pandas
        file = 'Grocery.xlsx'
        xl = pd.ExcelFile(file)
        
        # Generate DataFrame from imported spreadsheet and print
        df1 = xl.parse('Page 1')
        print (df1)

        #regenerate list of items in inventory
        list1 = [""]*(sheet_1.max_row)
        for t in range(0, sheet_1.max_row):
            list1[t] = sheet_1.cell(row = t+1, column = 1).value
            list2 = filter(lambda a: a != "Item", list1)
            list3 = ['{} '.format(elem) for elem in list2]
        print list3
        
        rowselect = list3.index(item)
        row4search = rowselect+2 #find in which row item exists
        print row4search


        code=sheet_1.cell(row=row4search, column=2).value #Fill in data to blank arrays from spreadsheet
        cost=sheet_1.cell(row=row4search, column=3).value
        aisle=sheet_1.cell(row=row4search, column=4).value
        bay=sheet_1.cell(row=row4search, column=5).value

        # Full info display
        print (sheet_1.cell(row=row4search, column=1).value)
        print ('Barcode {}'.format(code)) #Display desired item barcode
        print ('Price ${}'.format(cost)) #Display desired item cost
        print ('Aisle {}'.format(aisle)) #Display desired item isle
        print ('Bay {}'.format(bay)) #Display desired item bay

        # CART LIST
        def addCartList():
            shoppingCartButton = tk.Button(rightFrame1, text = item, bg = 'thistle')
            shoppingCartButton.pack(fill = tk.X)
            runningTot = runningTot+cost
        
        # SHOPPING LIST
        addCartButton = tk.Button(leftFrame1, text = 'ADD '+item+' $'+str(cost)+' Aisle '+str(aisle),
                                  command = addCartList, bg = 'plum')
        addCartButton.pack(fill = tk.X)
            
    def on_keyrelease(self, event):
        # get text from entry
        value = event.widget.get()
        value = value.strip().lower()
##        value = self.widget.get()
##        value = value.strip().lower()
        # get data from test_list
        if value == '':
            data = []
        else:
            data = []
            for item in self.test_list3:
                if value in item.lower():
                    data.append(item)                
        # update data in listbox
        listbox_update(data)
        
    def listbox_update(data):
        # delete previous data
        listbox.delete(0, 'end')
        # sorting data 
        data = sorted(data, key=str.lower)
        # put new data
        for item in data:
            listbox.insert('end', item)
        
    def gsearch(self):
        #keyboard()
        self.searchButton.destroy()
        # Search Entry
        entry = tk.Entry(self.centerFrame1)
        entry.pack(side = tk.TOP)
        entry.bind('<KeyRelease>', self.on_keyrelease)

        listbox = tk.Listbox(self.centerFrame1)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', self.on_select)
        listbox_update([])

    #def keyboard(self):

    

def main():
    root = tk.Tk()
    app = GUI(root)
    root.title("ElSA")
    screenWidth = root.winfo_screenwidth()
    screenHeight = root.winfo_screenheight()
    x = (screenWidth/2) - (screenWidth/2)
    y = (screenHeight/2) - (screenHeight/2)
    root.geometry('%dx%d+%d+%d' % (screenWidth, screenHeight, x, y))
    root.resizable(width = False, height = False)
    root.mainloop()

if __name__ == '__main__':
    main()
    
