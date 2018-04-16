#import tkinter in a universal way that would work for both python2 and python3
try:
    #Python 2
    import Tkinter as tk
except ImportError:
    #Python 3
    import tkinter as tk
#from Tkinter import *
import tkFont
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import os
import time



#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
### GUI definitions ###
root = tk.Tk()
root.title("ElSA")
myFont = tkFont.Font(family = 'Helvetica', size = 12, weight = "bold")

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

class Boom(tk.Frame):
    #global listbox
    global test_list3
    global listbox_update
    global sheet_1
    global itemlist
    global pricelist
    global on_keyrelease
    global on_select
    global topFrame1
    global leftFrame1
    global rightFrame1
    global centerFrame1
    global searchButton
    global change
    global screenWidth
    global screenHeight
##    global itemMaster
##    global priceMaster

    screenWidth = root.winfo_screenwidth()
    screenHeight = root.winfo_screenheight()
    x = (screenWidth/2) - (screenWidth/2)
    y = (screenHeight/2) - (screenHeight/2)
        
    root.geometry('%dx%d+%d+%d' % (screenWidth, screenHeight, x, y))
    root.resizable(width = False, height = False)

    print screenWidth
    print screenHeight
    
    # create all of the main containers
    topFrame1 = tk.Frame(root, bg='lightskyblue') #topFrame1 = tk.Frame(root, bg='lightskyblue', width = 1920, height = 200, padx = 5, pady = 5)
    topFrame2 = tk.Frame(root, bg='lightskyblue')
    topFrame3 = tk.Frame(root, bg='lightskyblue')
    leftFrame1 = tk.Frame(root, bg='blue')
    centerFrame1 = tk.Frame(root, bg='deepskyblue')
    rightFrame1 = tk.Frame(root, bg='darkblue')
    bottomFrame1 = tk.Frame(root, bg='magenta')

    # layout all of the main containers
    leftFrame1.columnconfigure(0, weight = 1, uniform = 'same')
    centerFrame1.columnconfigure(1, weight = 2, uniform = 'same')
    rightFrame1.columnconfigure(2, weight = 1, uniform = 'same')
    root.grid_rowconfigure(0, weight = 1, uniform = 'same1')
    root.grid_rowconfigure(1, weight = 7, uniform = 'same1')
    root.grid_rowconfigure(2, weight = 2, uniform = 'same1')
    root.grid_columnconfigure(0, weight = 1, uniform = 'same2')
    root.grid_columnconfigure(1, weight = 1, uniform = 'same2')
    root.grid_columnconfigure(2, weight = 1, uniform = 'same2')
    topFrame1.grid_rowconfigure(0, weight = 1, uniform = 'same3')
    topFrame2.grid_rowconfigure(1, weight = 1, uniform = 'same3')
    topFrame3.grid_rowconfigure(2, weight = 1, uniform = 'same3')

    topFrame1.grid(row=0, column = 0, sticky="nsew")
    topFrame2.grid(row=0, column = 1, sticky="nsew")
    topFrame3.grid(row=0, column = 2, sticky="nsew")
    leftFrame1.grid(row=1, column = 0, sticky="nsew")
    centerFrame1.grid(row=1, column = 1, sticky="nsew")
    rightFrame1.grid(row=1, column = 2, sticky="nsew")
    bottomFrame1.grid(row=2, columnspan = 3, sticky="nsew")

    bottomFrameWidth = bottomFrame1.winfo_width()
    print bottomFrameWidth
    
    itemMaster = ['']
    priceMaster = ['']

    
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
### Event Functions ###
    def close():
        root.destroy() #close GUI
    def listbox_update(data):
        # delete previous data
        listbox.delete(0, 'end')
        # sorting data 
        data = sorted(data, key=str.lower)
        # put new data
        for item in data:
            listbox.insert('end', item)
    def on_keyrelease(event):
        # get text from entry
        value = event.widget.get()
        value = value.strip().lower()
        # get data from test_list
        if value == '':
            data = []
        else:
            data = []
            for item in test_list3:
                if value in item.lower():
                    data.append(item)                
        # update data in listbox
        listbox_update(data)
    def listbox_update(data):
        # delete previous data
        listbox.delete(0, 'end')
        # sorting data 
        data = sorted(data, key=str.lower)
        # put new data
        for item in data:
            listbox.insert('end', item)
    def gsearch():
        searchButton.destroy()
        # Search Entry
        #entry = tk.Entry(topFrame1)
        entry = tk.Entry(centerFrame1)
        entry.pack(side =tk.TOP)
        entry.bind('<KeyRelease>', on_keyrelease)


        global listbox
        #listbox = tk.Listbox(topFrame1)
        listbox = tk.Listbox(centerFrame1)
        listbox.pack()
        listbox.bind('<<ListboxSelect>>', on_select)
        listbox_update([])
        import keyboard_import
    
    def on_select(event):
    # Takes user selection and removes price information in order to use selected item to search for other info
    #______________________________________________________________________________________________________________#
        grabitem = event.widget.get(event.widget.curselection()) #the initial selection of the user
        for t in range(0, sheet_1.max_row):
            itemlist[t] = sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1 describing item
            itemlist2 = ['{} '.format(elem) for elem in itemlist] #Formats to remove 'u' type from string

            pricelist[t] = ('   |${}|'.format(sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3 describing price
            pricelist2 = ['{} '.format(elem) for elem in pricelist] #Formats to remove 'u' type from string
        
        test_list = map(str.__add__, itemlist2, pricelist2) #combines lists into one in order to find index within list of selected item
        test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #clean up list
        print test_list3

        grabitemindex = test_list3.index(grabitem)+1 #finds index of selected item
        pricelength = len(pricelist2[grabitemindex]) #finds length of displayed price info in order to determine how long of a string to delete

        item = grabitem[:-pricelength] #deletes price info leaving only item title
        
##        for d in range(0, k):
##            if itemMaster[d] == '':
##                itemMaster[d] == item
##                itemMaster.append('')
##                print itemMaster
##                break
##
##        for p in range(len(priceMaster)):
##            if priceMaster[p] == '':
##                priceMaster[p] == item
##                priceMaster.append('')
##                print priceMaster
##                break
        
    #______________________________________________________________________________________________________________#

    #Below runs item through Pandas
        file = 'Grocery.xlsx'
        xl = pd.ExcelFile(file)
        
        # Generate DataFrame from imported spreadsheet and print
        df1 = xl.parse('Page 1')
        print (df1)

        #regenerate list of items in inventory
        list1 = [""]*(sheet_1.max_row)
        for t in range(0, sheet_1.max_row):
            list1[t] = sheet_1.cell(row = t+1, column = 1).value
            list2 = filter(lambda a: a != "Item", list1)
            list3 = ['{} '.format(elem) for elem in list2]
        print list3
        
        rowselect = list3.index(item)
        row4search = rowselect+2 #find in which row item exists
        print row4search


        code=sheet_1.cell(row=row4search, column=2).value #Fill in data to blank arrays from spreadsheet
        cost=sheet_1.cell(row=row4search, column=3).value
        aisle=sheet_1.cell(row=row4search, column=4).value
        bay=sheet_1.cell(row=row4search, column=5).value

        # Full info display
        print (sheet_1.cell(row=row4search, column=1).value)
        print ('Barcode {}'.format(code)) #Display desired item barcode
        print ('Price ${}'.format(cost)) #Display desired item cost
        print ('Aisle {}'.format(aisle)) #Display desired item isle
        print ('Bay {}'.format(bay)) #Display desired item bay

        # CART LIST
        def addCartList():
            shoppingCartButton = tk.Button(rightFrame1, text = item, bg = 'thistle')
            #shoppingCartButton.grid(column = 1)
            shoppingCartButton.pack(fill = tk.X)
            runningTot = runningTot+cost
        
        # SHOPPING LIST
        addCartButton = tk.Button(leftFrame1, text = 'ADD '+item+' $'+str(cost)+' Aisle '+str(aisle), command = addCartList, bg = 'plum')
        #addCartButton.grid(column = 1)
        addCartButton.pack(fill = tk.X)
        
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
# --- Main --- #

    #Loads workbook to look for items in inventory 
    wb = load_workbook('Grocery.xlsx')
    sheet_1 = wb['Page 1']
    itemlist = [""]*(sheet_1.max_row)
    pricelist = [""]*(sheet_1.max_row)
    for t in range(0, sheet_1.max_row):
        itemlist[t] = sheet_1.cell(row = t+1, column = 1).value #Pulls all strings from column 1
        itemlist2 = ['{} '.format(elem) for elem in itemlist] #Formats to remove 'u' type from string

        pricelist[t] = ('   |${}|'.format(sheet_1.cell(row = t+1, column = 3).value)) #Pulls all strings from column 3
        pricelist2 = ['{} '.format(elem) for elem in pricelist] #Formats to remove 'u' type from string
    
    test_list = map(str.__add__, itemlist2, pricelist2)
    
    test_list3 = filter(lambda a: a != 'Item    |$Cost| ', test_list) #Combines item with price info for user
    print test_list3

    # Shopping Cart Label
    shoppingCartLabel = tk.Label(rightFrame1, text = "Your Shopping Cart:", bg = 'orchid', fg = 'floralwhite', font = myFont)
    #shoppingCartLabel.grid(column = 1, row = 1, rowspan = 2, padx = 10)
    shoppingCartLabel.pack(fill = tk.X)

    # Total Cost Label
    totalCostLabel = tk.Label(rightFrame1, text = "Total Cost: ", bg = 'cornflowerblue', fg = 'lavenderblush', font = myFont)
    totalCostLabel.pack(fill = tk.X, side = tk.BOTTOM)

    # Shopping List Label
    addCartLabel = tk.Label(leftFrame1, text = "Shopping List\nClick to add items to cart: ", bg = 'cornflowerblue', fg = 'lavenderblush', font = myFont)
    #addCartLabel.grid(column = 1, row = 1, padx = 10)
    addCartLabel.pack(fill = tk.X)
    
    # Search button
    searchButton = tk.Button(centerFrame1, text = 'Search', bg = 'mediumslateblue', fg = 'snow', font = myFont, command = gsearch, height = 9)
    #searchButton = tk.Button(topFrame1, text = 'Search', bg = 'mediumslateblue', fg = 'snow', font = myFont, command = gsearch, height = 5)
    #searchButton = tk.Button(topFrame1, text = 'Search', bg = 'mediumslateblue', fg = 'snow', font = myFont, command = change)
    #searchPicture = tk.PhotoImage(file = "Search.png")
    #searchButton.config(image = searchPicture)
    searchButton.pack(fill = tk.X)

    # Store Logo
    logoLabel = tk.Label(topFrame2, text = 'Place logo here!', font = myFont, height = 10, width = 30)
    logoLabel.pack()

    # Ad Button
    adButton = tk.Button(bottomFrame1, text = 'Place ad here!', font = myFont, height = 96, width = 800)
    photoMilk = tk.PhotoImage(file = "HorizonAd.png")
    adButton.config(image = photoMilk)
    adButton.pack()

    # Recipe Builder Button
    recipeButton = tk.Button(centerFrame1, text = 'Place recipes here!', font = myFont, height = 10)
    recipeButton.pack(side = tk.BOTTOM, fill = tk.X)


#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#
#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#


#Display the menu
frame01 = Boom()
frame01.mainloop()
