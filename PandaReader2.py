import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np

#vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv#
#Item look-up initial set-up
#__________________________________________________________________________#

#Import desired inventory file and save as a variable
file = 'Grocery.xlsx'
xl = pd.ExcelFile(file)

#Display all sheets that exist in spreadsheet file
print(xl.sheet_names)

#Generate DataFrame from imported spreadsheet and print
df1 = xl.parse('Page 1')
print df1

shoplist = ["" for x in range(1)]

item = "Start"
while item != "Done":
    item = input("Add item or type Done: ")

        #except ValueError:
            #print("Sorry, invalid input. Try adding another")
    if item != "Done":
        if df1['Item'].str.match(item).any():
            for j in range(len(shoplist)):
                if shoplist[j] == "":
                    shoplist[j] = item
                    shoplist.append("")
                    break
        else:
            print("Item not found. Please add a new Item.")
    else:
        break
print shoplist

    
# Use to test location of specific item-->> print(df1.loc[df1['Item'] == 'Nuts'])

#Find row in which desired item exists
for p in range(len(shoplist)-1):
 #  rowselect = (df1.index[df1['Item'] == shoplist[p]])
    rowselect = (df1.index[df1['Item'].str.match(shoplist[p])])
    row4search = rowselect+1 #Variable used to search for numerical information in sheet
    #Display int64 values
    #print(rowselect)
    #print(row4search)
    #Display parsed row of item info
    #print(df1.loc[rowselect])
#__________________________________________________________________________#
#Initial set-up complete
#^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^#


#vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv#
#Re-load spreadsheet using openpyxl in order to extract individual item info
#__________________________________________________________________________#
    wb = load_workbook('Grocery.xlsx') #Loads spreadsheet
    sheet_1 = wb.get_sheet_by_name('Page 1') #Extracts desired sheet


    code = np.zeros(sheet_1.max_row) #Create blank array based on size of inventory list
    cost = np.zeros(sheet_1.max_row)
    aisle = np.zeros(sheet_1.max_row)
    bay = np.zeros(sheet_1.max_row)
 
    for i in range(1,sheet_1.max_row):
        code[i]=sheet_1.cell(row=i+1, column=2).value #Fill in data to blank arrays from spreadsheet
        cost[i]=sheet_1.cell(row=i+1, column=3).value
        aisle[i]=sheet_1.cell(row=i+1, column=4).value
        bay[i]=sheet_1.cell(row=i+1, column=5).value

#Display lists 
    #print code
    #print cost
    #print aisle
    #print bay

#Full info display
    print shoplist[p]
    print ('Barcode', code[row4search]) #Display desired item barcode
    print ('Price', cost[row4search]) #Display desired item cost
    print ('Aisle', aisle[row4search]) #Display desired item isle
    print ('Bay', bay[row4search]) #Display desired item bay
#__________________________________________________________________________#
#^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^#
