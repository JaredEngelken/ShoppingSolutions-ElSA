from Tkinter import *
import tkFont
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import os
import time

global clear

### GUI definitions ###
root = Tk()
root.title("ElSA")
myFont = tkFont.Font(family = 'Helvetica', size = 12, weight = "bold")

class Boom(Frame):
    ### Event Functions ###
    def close():
        root.destroy() #close GUI
    def clear():
        if 'dneLabel' in globals():
            dneLabel.destroy()
        elif 'itemButton' in globals():
            itemButton.destroy()
    def pandas():
        #def clear():
         #   if 'itemLabel' in globals():
         #       itemLabel.destroy()
         #   elif 'itemButton' in globals():
         #       itemButton.destroy()
        # Import desired inventory file and save as a variable
        file = 'Grocery.xlsx'
        xl = pd.ExcelFile(file)

        # Display all sheets that exist in spreadsheet file
        print(xl.sheet_names)

        # Generate DataFrame from imported spreadsheet and print
        df1 = xl.parse('Page 1')
        print (df1)
        shoplist = ["" for x in range(1)]

        # Get the item from the search box and make a button with searched item if found
        item = ment.get()

          
        if item != "Done":
            global itemButton
            global itemLabel
            if df1['Item'].str.match(item).any():
                for j in range(len(shoplist)):
                    if shoplist[j] == "":
                        shoplist[j] = item
                        shoplist.append("")
                        print (shoplist)
                                # Use to test location of specific item-->> print(df1.loc[df1['Item'] == 'Nuts'])

                            # Find row in which desired item exists
                        for p in range(len(shoplist)-1):
                            row4search = ((df1.index[df1['Item'].str.contains(shoplist[p])])+1)

                            wb = load_workbook('Grocery.xlsx') #Loads spreadsheet
                            sheet_1 = wb['Page 1'] #Extracts desired sheet
                            for totMatch in range(len(row4search)):
                                code = np.zeros(sheet_1.max_row) #Create blank array based on size of inventory list
                                cost = np.zeros(sheet_1.max_row)
                                aisle = np.zeros(sheet_1.max_row)
                                bay = np.zeros(sheet_1.max_row)
                             
                                for i in range(1,sheet_1.max_row):
                                    code[i]=sheet_1.cell(row=i+1, column=2).value #Fill in data to blank arrays from spreadsheet
                                    cost[i]=sheet_1.cell(row=i+1, column=3).value
                                    aisle[i]=sheet_1.cell(row=i+1, column=4).value
                                    bay[i]=sheet_1.cell(row=i+1, column=5).value                                
                                        
                            # Full info display
                                print (sheet_1.cell(row=row4search[totMatch]+1, column=1).value)
                                print ('Barcode', code[row4search[totMatch]]) #Display desired item barcode
                                print ('Price', cost[row4search[totMatch]]) #Display desired item cost
                                print ('Aisle', aisle[row4search[totMatch]]) #Display desired item isle
                                print ('Bay', bay[row4search[totMatch]]) #Display desired item bay

                            # CART LIST
                                def addCartList():
                                    shoppingCartButton = Button(root, text = item)
                                    shoppingCartButton.grid(column = 4)
                                
                            # SHOPPING LIST
                                def addShopList():
                                    addCartButton = Button(root, text = item, command = addCartList)
                                    addCartButton.grid(column = 3) 
                                
                                #Add item to shoplist button with info (!!!ONLY DISPLAYS SEARCH NOT ACTUAL NAME OF ITEM!!!)
                                itemButton = Button(root, text = 'ADD '+item+' $'+str(cost[row4search[totMatch]])+' Aisle '+str(aisle[row4search[totMatch]]), command = addShopList)
                                itemButton.grid(column = 1, columnspan = 2)
                                # MAKE THIS THING DISAPEAR ON NEXT SEARCH!!!
                        break
            else:
                dneLabel = Label(root, text = 'Item not found.', font = myFont)
                dneLabel.grid(column = 1)
    # Clear Button
    itemClear = Button(root, text = 'Clear item', command = clear, fg = 'red', bg = 'white', font = myFont)
    itemClear.grid(column = 2, row = 1)

    # Shopping Cart Label
    shoppingCartLabel = Label(root, text = "Your Shopping Cart:", bg = 'green', fg = 'black', font = myFont)
    shoppingCartLabel.grid(column = 4, row = 1)

    # Shopping List Label
    addCartLabel = Label(root, text = "Shopping List\nClick to add items to cart: ", bg = 'white', fg = 'black', font = myFont)
    addCartLabel.grid(column = 3, row = 1)

    #Entry Box and Search Button
    global ment
    ment = StringVar()
    search = Button(root, text = 'Search', command = pandas, fg = 'white', bg = 'blue', font = myFont)
    search.grid(column = 1, row = 1)
    mEntry = Entry(root, textvariable = ment)
    mEntry.grid(column = 1, columnspan = 2, row = 2)

#Display the menu
frame01 = Boom()
#root.mainloop() #loops forever
frame01.mainloop()
